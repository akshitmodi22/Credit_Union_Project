# ═══════════════════════════════════════════════════════════════
# reports/excel_builder.py — Excel report generation
# ═══════════════════════════════════════════════════════════════
import pandas as pd
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


def build_excel_with_validation(
    df         : pd.DataFrame,
    filename   : str,
    validation : dict
):
    """
    Saves Excel file with 2 sheets:
    Sheet 1 → Member Data (with 5 SHAP reasons)
    Sheet 2 → Validation results
    """

    # ── Reorder columns for better readability ─────────────────
    # Priority columns to show first
    priority_cols = [
        'member_id',
        'branch_assignment',
        'county',
        'age_band',
        'income_band',
        'life_stage',
        'employment_status',
        'tenure_years',
        'credit_score_band',
    ]

    # Model score columns
    score_cols = [
        'attrition_probability', 'attrition_tier',
        'loan_offer_score',      'loan_offer_tier',
        'propensity_probability','propensity_tier',
        'predicted_status',
        'predicted_label',
        'actual_label',
    ]

    # SHAP reason columns — top 5
    shap_cols = [f'shap_reason_{i}' for i in range(1, 6)]

    # Build final column order
    ordered_cols = []
    for col in priority_cols + score_cols + shap_cols:
        if col in df.columns:
            ordered_cols.append(col)

    # Add any remaining columns not already included
    remaining = [c for c in df.columns if c not in ordered_cols]
    ordered_cols += remaining

    df_ordered = df[ordered_cols].copy()

    # ── Write to Excel ─────────────────────────────────────────
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        # Sheet 1 — Member Data
        df_ordered.to_excel(
            writer,
            sheet_name = 'Member Data',
            index      = False
        )

        # ── Style Sheet 1 ──────────────────────────────────────
        ws = writer.sheets['Member Data']

        # Header style
        header_fill = PatternFill(
            start_color='1E3A5F',
            end_color='1E3A5F',
            fill_type='solid'
        )
        header_font = Font(
            color='FFFFFF',
            bold=True,
            size=10
        )
        header_align = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        for col_idx, col_name in enumerate(df_ordered.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align

        # Color tier cells
        tier_colors = {
            'High'  : 'FFE5E5',  # light red
            'Medium': 'FFF3CD',  # light amber
            'Low'   : 'E8F5E9',  # light green
        }

        tier_col_names = [
            'attrition_tier', 'loan_offer_tier', 'propensity_tier'
        ]

        for col_idx, col_name in enumerate(df_ordered.columns, 1):
            if col_name in tier_col_names:
                for row_idx in range(2, len(df_ordered) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val  = str(cell.value) if cell.value else ''
                    if val in tier_colors:
                        cell.fill = PatternFill(
                            start_color=tier_colors[val],
                            end_color=tier_colors[val],
                            fill_type='solid'
                        )
                        cell.font = Font(bold=True, size=9)

        # SHAP reason column styling
        shap_col_indices = [
            col_idx for col_idx, col_name
            in enumerate(df_ordered.columns, 1)
            if col_name.startswith('shap_reason_')
        ]
        for col_idx in shap_col_indices:
            for row_idx in range(2, len(df_ordered) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font      = Font(size=8, color='555555')
                cell.alignment = Alignment(wrap_text=True)

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df_ordered.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col_name.startswith('shap_reason_'):
                ws.column_dimensions[col_letter].width = 35
            elif col_name in ['member_id', 'branch_assignment']:
                ws.column_dimensions[col_letter].width = 18
            elif col_name in score_cols:
                ws.column_dimensions[col_letter].width = 16
            else:
                ws.column_dimensions[col_letter].width = 14

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Row height for header
        ws.row_dimensions[1].height = 30

        # ── Sheet 2 — Validation ───────────────────────────────
        val_rows = []
        for check_name, check_val in validation.items():
            if isinstance(check_val, dict):
                passed = check_val.get('passed', True)
                val_rows.append({
                    'Check'   : check_name,
                    'Details' : str(check_val),
                    'Status'  : '✓ PASS' if passed else '✗ FAIL'
                })
            else:
                val_rows.append({
                    'Check'   : check_name,
                    'Details' : str(check_val),
                    'Status'  : '✓ PASS' if check_val else '✗ FAIL'
                })

        val_df = pd.DataFrame(val_rows)
        val_df.to_excel(
            writer,
            sheet_name = 'Validation',
            index      = False
        )

        # Style validation sheet
        ws2 = writer.sheets['Validation']
        for col_idx in range(1, 4):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = PatternFill(
                start_color='1E3A5F',
                end_color='1E3A5F',
                fill_type='solid'
            )
            cell.font = Font(color='FFFFFF', bold=True)

        # Color pass/fail rows
        for row_idx in range(2, len(val_rows) + 2):
            status_cell = ws2.cell(row=row_idx, column=3)
            if status_cell.value and '✓' in str(status_cell.value):
                for col_idx in range(1, 4):
                    ws2.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'
                    )
            else:
                for col_idx in range(1, 4):
                    ws2.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color='FFE5E5', end_color='FFE5E5', fill_type='solid'
                    )

        # Auto width validation sheet
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 12